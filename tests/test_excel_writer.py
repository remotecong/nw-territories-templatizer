from pathlib import Path
import openpyxl
from src.excel_writer import create_territory_workbook

def test_create_territory_workbook(tmp_path: Path):
    territory = {"CategoryCode": "A", "Number": "12", "Area": "Bandon Trails"}
    street_data = {
        "East 94th Pl": [
            {"address": "2935 East 94th Pl #101", "name": "John Doe", "phone": "555-1111"},
            {"address": "2935 East 94th Pl #102", "name": "", "phone": ""},
        ]
    }
    out_file = tmp_path / "A12.xlsx"
    create_territory_workbook(territory, street_data, out_file)

    assert out_file.exists()
    wb = openpyxl.load_workbook(out_file)
    assert "East 94th Pl" in wb.sheetnames
    ws = wb["East 94th Pl"]

    # Check Row 1 Title
    assert ws["A1"].value == "A12 - Bandon Trails"
    # Check Row 2 Notes
    assert ws["A2"].value == "Notes:"
    # Check Row 3 Headers
    assert ws["A3"].value == "Address"
    assert ws["B3"].value == "Name"
    assert ws["C3"].value == "Phone"
    # Check Row 4 Data
    assert ws["A4"].value == "2935 East 94th Pl #101"
    assert ws["B4"].value == "John Doe"
    assert ws["C4"].value == "555-1111"

def test_create_territory_workbook_empty_area(tmp_path: Path):
    territory = {"CategoryCode": "R", "Number": "5", "Area": ""}
    street_data = {"Main St": [{"address": "100 Main St", "name": "", "phone": ""}]}
    out_file = tmp_path / "R5.xlsx"
    create_territory_workbook(territory, street_data, out_file)

    wb = openpyxl.load_workbook(out_file)
    ws = wb["Main St"]
    assert ws["A1"].value == "R5"

def test_create_territory_workbook_none_string_area(tmp_path: Path):
    territory = {"CategoryCode": "R", "Number": "5", "Area": "None"}
    street_data = {"Main St": [{"address": "100 Main St", "name": "", "phone": ""}]}
    out_file = tmp_path / "R5.xlsx"
    create_territory_workbook(territory, street_data, out_file)

    wb = openpyxl.load_workbook(out_file)
    ws = wb["Main St"]
    assert ws["A1"].value == "R5"
