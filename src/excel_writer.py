from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def clean_sheet_title(title: str, existing_titles: set[str]) -> str:
    r"""Excel sheet names are max 31 chars and cannot contain: \ / ? * : [ ]"""
    invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
    cleaned = title
    for ch in invalid_chars:
        cleaned = cleaned.replace(ch, '')
    cleaned = cleaned[:31].strip()
    if not cleaned:
        cleaned = "Street"

    base = cleaned
    counter = 1
    while cleaned in existing_titles:
        suffix = f"_{counter}"
        cleaned = f"{base[:31 - len(suffix)]}{suffix}"
        counter += 1
    existing_titles.add(cleaned)
    return cleaned

def create_territory_workbook(territory: dict, street_addresses: dict[str, list[dict]], output_path: Path) -> Path:
    """Generates an .xlsx workbook for a territory."""
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    cat_code = territory.get("CategoryCode", "").strip()
    num = territory.get("Number", "").strip()
    area = territory.get("Area", "").strip()
    if area.lower() in ("none", "null", "nil"):
        area = ""

    title_text = f"{cat_code}{num} - {area}" if area else f"{cat_code}{num}"

    bold_font = Font(bold=True)
    existing_titles = set()

    for street_name, records in street_addresses.items():
        sheet_title = clean_sheet_title(street_name, existing_titles)
        ws = wb.create_sheet(title=sheet_title)

        # Row 1: Title
        ws.merge_cells("A1:C1")
        ws["A1"] = title_text
        ws["A1"].font = Font(bold=True, size=12)
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

        # Row 2: Notes
        ws.merge_cells("A2:C2")
        ws["A2"] = "Notes:"
        ws["A2"].font = Font(italic=True)
        ws["A2"].alignment = Alignment(horizontal="left", vertical="top")
        ws.row_dimensions[2].height = 40  # Allow some height for notes

        # Row 3: Headers
        headers = ["Address", "Name", "Phone"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = bold_font

        # Row 4+: Data
        for row_idx, item in enumerate(records, start=4):
            ws.cell(row=row_idx, column=1, value=item.get("address", ""))
            ws.cell(row=row_idx, column=2, value=item.get("name", ""))
            ws.cell(row=row_idx, column=3, value=item.get("phone", ""))

        # Auto-fit columns
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                # Skip merged rows (1 and 2) for width calculation
                if cell.row in (1, 2):
                    continue
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
